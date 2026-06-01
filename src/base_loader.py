"""
Carregamento de Cofre_atestados_brasul.xlsx (abas Atestados + Referência).
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

import pandas as pd

from config.settings import ABA_ATESTADOS, ABA_REFERENCIA, COLUNAS_COFRE
from utils.tipos import normalizar_tipo


def _c7(cod) -> str:
    return re.sub(r'\D', '', str(cod or ''))


def _norm_txt(txt: str) -> str:
    t = unicodedata.normalize('NFD', str(txt or ''))
    return ''.join(c for c in t if unicodedata.category(c) != 'Mn').upper()


def _norm_arquivo(nome: str) -> str:
    """Ignora diferença de espaço/underscore (ex.: Aldeia Capoeirao vs AldeiaCapoeirao)."""
    s = str(nome or '').strip().lower()
    return re.sub(r'[\s_\-]+', '', s)


def formatar_codigo(c7: str) -> str:
    return f"{c7[:2]}.{c7[2:4]}.{c7[4:]}"


@dataclass
class BaseContext:
    """Índices carregados a partir da planilha de atestados."""

    por_cod: dict = field(default_factory=dict)           # c7 -> (desc, un)
    por_desc: list = field(default_factory=list)          # [(c7, DESC_UPPER), ...]
    por_arquivo: dict = field(default_factory=dict)       # pdf -> {c7 -> item}
    por_obra: dict = field(default_factory=dict)          # obra_norm -> {c7 -> item}
    valid_codigos: set = field(default_factory=set)
    nome_obra_por_arquivo: dict = field(default_factory=dict)  # pdf -> OBRA


def _registrar_item(ctx: BaseContext, c7: str, desc: str, un: str,
                    obra: str = '', arquivo: str = '', tipo: str = '',
                    data_inicio=None, data_emissao=None) -> None:
    if len(c7) != 7:
        return

    desc = str(desc or '').strip()
    un = str(un or '').strip()
    tipo = normalizar_tipo(tipo)

    ctx.por_cod[c7] = (desc, un)
    ctx.valid_codigos.add(c7)
    if desc:
        ctx.por_desc.append((c7, desc.upper()))

    item = {
        'codigo': formatar_codigo(c7),
        'descricao': desc,
        'unidade': un,
        'tipo': tipo,
        'obra': str(obra or '').strip(),
        'data_inicio': data_inicio,
        'data_emissao': data_emissao,
    }

    arq_key = _norm_arquivo(arquivo)
    if arq_key:
        ctx.por_arquivo.setdefault(arq_key, {})[c7] = item
        if obra:
            ctx.nome_obra_por_arquivo[arq_key] = str(obra).strip()

    obra_key = _norm_txt(obra)
    if obra_key:
        ctx.por_obra.setdefault(obra_key, {})[c7] = item


def carregar_base(caminho: Path) -> BaseContext:
    """
    Carrega Cofre_atestados_brasul.xlsx (input).

    Abas esperadas:
      - Atestados de Obras: OBRA, Obra_Arquivo, Tipo, Cod, Desc, UN, datas
      - Base de Referência: Cod, Desc, UN (sem linha de cabeçalho)
    """
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha de atestados não encontrada:\n  {caminho}")

    ctx = BaseContext()
    wb = openpyxl.load_workbook(str(caminho), data_only=True, read_only=True)

    # ── Aba de referência global (códigos FDE) ──
    if ABA_REFERENCIA in wb.sheetnames:
        ws = wb[ABA_REFERENCIA]
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            c7 = _c7(row[0])
            if len(c7) != 7:
                continue
            _registrar_item(ctx, c7, row[1] if len(row) > 1 else '',
                            row[2] if len(row) > 2 else '')

    # ── Aba por obra / PDF (dados validados) ──
    if ABA_ATESTADOS in wb.sheetnames:
        ws = wb[ABA_ATESTADOS]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                if row and str(row[0] or '').upper() == 'OBRA':
                    continue
            if not row or not row[3]:
                continue
            c7 = _c7(row[3])
            if len(c7) != 7:
                continue
            _registrar_item(
                ctx, c7,
                row[4] if len(row) > 4 else '',
                row[5] if len(row) > 5 else '',
                obra=row[0] if len(row) > 0 else '',
                arquivo=row[1] if len(row) > 1 else '',
                tipo=row[2] if len(row) > 2 else '',
                data_inicio=row[6] if len(row) > 6 else None,
                data_emissao=row[7] if len(row) > 7 else None,
            )

    wb.close()

    # Remove duplicatas em por_desc mantendo o último
    seen = set()
    por_desc_unicos = []
    for c7, desc in reversed(ctx.por_desc):
        if c7 not in seen:
            seen.add(c7)
            por_desc_unicos.append((c7, desc))
    ctx.por_desc = list(reversed(por_desc_unicos))

    return ctx


def cofre_precisa_atualizar(origem: Path, destino: Path) -> bool:
    """True se o cofre não existe ou a base foi alterada depois dele."""
    if not origem.exists():
        return False
    if not destino.exists():
        return True
    return origem.stat().st_mtime > destino.stat().st_mtime


def sincronizar_cofre_se_necessario(origem: Path, destino: Path) -> dict | None:
    if not cofre_precisa_atualizar(origem, destino):
        return None
    return sincronizar_cofre_da_base(origem, destino)


def sincronizar_cofre_da_base(origem: Path, destino: Path) -> dict:
    """
    Copia a aba Atestados para Cofre_Brasul.xlsx (consulta rápida na interface).
    """
    if not origem.exists():
        raise FileNotFoundError(f"Base perfeita não encontrada:\n  {origem}")

    df = pd.read_excel(origem, sheet_name=ABA_ATESTADOS).fillna('')

    for col in COLUNAS_COFRE:
        if col not in df.columns:
            df[col] = ''

    df = df[COLUNAS_COFRE]
    df["Tipo"] = df["Tipo"].apply(normalizar_tipo)
    destino.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(destino, index=False)

    return {
        'linhas': len(df),
        'obras': int(df['OBRA'].nunique()) if 'OBRA' in df.columns else 0,
        'pdfs': int(df['Obra_Arquivo'].nunique()) if 'Obra_Arquivo' in df.columns else 0,
        'codigos_unicos': int(df['Cod'].astype(str).nunique()) if 'Cod' in df.columns else 0,
    }
