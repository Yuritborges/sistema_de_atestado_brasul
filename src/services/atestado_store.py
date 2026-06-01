"""
Grava novos itens em Cofre_atestados_brasul.xlsx (aba Atestados de Obras).
"""

from pathlib import Path

import openpyxl
import pandas as pd

from config.settings import ABA_ATESTADOS, COLUNAS_COFRE
from utils.excel_io import PlanilhaEmUsoError, garantir_excel_livre, salvar_workbook
from utils.tipos import normalizar_tipo


def adicionar_itens_atestado(caminho_base: Path, itens: list[dict]) -> int:
    if not itens:
        return 0

    if not caminho_base.exists():
        raise FileNotFoundError(f"Planilha não encontrada:\n  {caminho_base}")

    garantir_excel_livre(caminho_base)

    wb = openpyxl.load_workbook(str(caminho_base))
    if ABA_ATESTADOS not in wb.sheetnames:
        wb.create_sheet(ABA_ATESTADOS)
        ws = wb[ABA_ATESTADOS]
        ws.append(COLUNAS_COFRE)
    else:
        ws = wb[ABA_ATESTADOS]

    for item in itens:
        row = []
        for col in COLUNAS_COFRE:
            val = item.get(col, "")
            if col == "Tipo":
                val = normalizar_tipo(val)
            elif col in ("OBRA", "Desc", "Obra_Arquivo"):
                val = str(val or "").strip().upper()
            row.append(val if val is not None else "")
        ws.append(row)

    salvar_workbook(wb, caminho_base, fazer_backup=True)
    return len(itens)


def ler_atestados_df(caminho_base: Path) -> pd.DataFrame:
    return pd.read_excel(caminho_base, sheet_name=ABA_ATESTADOS).fillna("")
