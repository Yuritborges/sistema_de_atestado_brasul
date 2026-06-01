"""
Grava novos códigos na aba Base de Referência.
"""

from pathlib import Path

import openpyxl

from config.settings import ABA_REFERENCIA
from services.referencia import _c7, formatar_codigo
from utils.excel_io import garantir_excel_livre, salvar_workbook


def adicionar_codigo_referencia(caminho: Path, codigo: str, descricao: str, unidade: str) -> str:
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada:\n  {caminho}")

    c7 = _c7(codigo)
    if len(c7) != 7:
        raise ValueError("Código FDE deve ter 7 dígitos (formato XX.XX.XXX).")

    cod_fmt = formatar_codigo(c7)
    descricao = str(descricao or "").strip().upper()
    unidade = str(unidade or "").strip().upper()

    if not descricao:
        raise ValueError("Informe a descrição do código.")

    garantir_excel_livre(caminho)

    wb = openpyxl.load_workbook(str(caminho))
    if ABA_REFERENCIA not in wb.sheetnames:
        wb.create_sheet(ABA_REFERENCIA)
    ws = wb[ABA_REFERENCIA]
    ws.append([cod_fmt, descricao, unidade])
    salvar_workbook(wb, caminho, fazer_backup=True)
    return cod_fmt
