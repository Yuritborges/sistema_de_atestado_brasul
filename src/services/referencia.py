"""
Catálogo da aba Base de Referência (código -> descrição + unidade).
"""

import re
from pathlib import Path

import openpyxl

from config.settings import ABA_REFERENCIA


def _c7(cod: str) -> str:
    return re.sub(r"\D", "", str(cod or ""))


def formatar_codigo(c7: str) -> str:
    if len(c7) != 7:
        return str(c7)
    return f"{c7[:2]}.{c7[2:4]}.{c7[4:]}"


def carregar_referencia(caminho: Path) -> dict:
    """
    Retorna {c7: {"codigo": "09.03.005", "descricao": "...", "unidade": "M"}}.
    """
    if not caminho.exists():
        return {}

    wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
    if ABA_REFERENCIA not in wb.sheetnames:
        wb.close()
        return {}

    catalogo = {}
    ws = wb[ABA_REFERENCIA]
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        c7 = _c7(row[0])
        if len(c7) != 7:
            continue
        desc = str(row[1] or "").strip()
        un = str(row[2] or "").strip()
        catalogo[c7] = {
            "codigo": formatar_codigo(c7),
            "descricao": desc,
            "unidade": un,
        }

    wb.close()
    return catalogo


def buscar_codigo(texto: str, catalogo: dict) -> dict | None:
    """Busca código digitado no catálogo (tolera pontos e espaços)."""
    c7 = _c7(texto)
    if len(c7) == 7 and c7 in catalogo:
        return catalogo[c7].copy()

    if not texto:
        return None

    t = str(texto).strip().upper()
    for item in catalogo.values():
        if item["codigo"].upper() == t:
            return item.copy()

    return None
